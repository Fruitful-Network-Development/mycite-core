"""Import BPW job rows from the operator's CRM spreadsheet into the
on-disk YAML store that backs the dashboard's Jobs tab.

Source : ``/srv/webapps/clients/brockspressurewashing.com/docs/Pressure_Washing_CRM.xlsx``
         (``Leads & Jobs`` sheet)
Target : ``/srv/webapps/mycite/fnd/private/utilities/tools/bpw-jobs/job.<date>.<slug>.yaml``

Strictly additive: a row is imported only when no YAML file already
exists at the canonical filename derived from ``(job_date,
customer_slug)``. Rows already represented on disk are left untouched —
the operator's hand-edits to existing YAMLs are preserved.

The xlsx column layout (header row):

    Date Contacted | Name | Phone | Address | Service Type | Lead Source |
    Status | Quote $ | Job Date | Paid? | Payment Method | Referred By |
    Date Confirm | Notes | Photos

Mapping conventions:

* ``Service Type`` → one or more ``tags[].type`` (canonical: house_wash,
  driveway, patio, front_walk, deck, fence, roof, gutter, commercial,
  other). Combined services like "Driveway + Patio" emit one tag per
  half.
* ``Lead Source`` → ``customer.lead_source`` (lowercase, underscored).
* ``Status`` → ``job.status`` (lowercase: completed | booked | cancelled
  | no_response).
* ``Payment Method`` → ``pricing.method`` (lowercase: cash | check |
  venmo | cashapp | paypal | card).
* ``Paid?`` "Yes" → ``pricing.paid: true``.
* ``Phone`` (float / string) is normalized to ``(NNN) NNN-NNNN`` when
  it contains 10 digits; otherwise rendered verbatim or dropped.

Usage:

    python -m MyCiteV2.scripts.bpw_jobs_import_from_xlsx --dry-run
    python -m MyCiteV2.scripts.bpw_jobs_import_from_xlsx
    python -m MyCiteV2.scripts.bpw_jobs_import_from_xlsx \\
        --xlsx <path> --jobs-root <path>
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date as _date
from datetime import datetime
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from MyCiteV2.instances._shared.runtime.utilities_extensions.bpw_jobs import (
    _DEFAULT_BPW_JOBS_ROOT,
    _filename_for,
    list_jobs,
    save_job,
)

DEFAULT_XLSX = Path(
    "/srv/webapps/clients/brockspressurewashing.com/docs/Pressure_Washing_CRM.xlsx"
)


# Canonical-key normalizers
_LEAD_SOURCE_MAP = {
    "facebook": "facebook",
    "flyer": "flyer",
    "flier": "flyer",
    "instagram": "instagram",
    "google": "google",
    "nextdoor": "nextdoor",
    "referral": "referral",
    "repeat customer": "repeat_customer",
    "door knocking": "door_knock",
    "door knock": "door_knock",
    "door hanger": "door_knock",
    "neighbor": "neighbor",
    "other": "other",
}

_STATUS_MAP = {
    "completed": "completed",
    "booked": "booked",
    "cancelled": "cancelled",
    "canceled": "cancelled",
    "no response": "no_response",
    "quoted": "booked",
    "lead": "no_response",
    "new": "no_response",
}

_PAYMENT_MAP = {
    "cash": "cash",
    "check": "check",
    "venmo": "venmo",
    "cash app": "cashapp",
    "cashapp": "cashapp",
    "paypal": "paypal",
    "card": "card",
    "credit card": "card",
}

# ``Service Type`` → list of canonical tag types.
_SERVICE_TAG_MAP = {
    "house wash": ["house_wash"],
    "housewash": ["house_wash"],
    "driveway": ["driveway"],
    "patio": ["patio"],
    "deck": ["deck"],
    "fence": ["fence"],
    "deck/fence": ["deck"],
    "deck/ fence": ["deck"],
    "roof": ["roof"],
    "gutter": ["gutter"],
    "gutter cleaning": ["gutter"],
    "commercial": ["commercial"],
    "other": ["other"],
}

_PHONE_DIGITS = re.compile(r"\D")
_SLUG_RE = re.compile(r"[^a-z0-9]+")


def _slugify(value: str) -> str:
    """Match _filename_for's customer slugifier exactly."""
    s = _SLUG_RE.sub(".", str(value or "").strip().lower()).strip(".")
    return s or "unknown"


def _existing_index(jobs_root: Path) -> set[tuple[str, str]]:
    """Build a set of (date_iso, normalized_name_key) tuples for every
    job YAML currently on disk.

    Filename conventions have varied over time (some files use a
    last-name slug, some a full-name slug), so we cannot dedupe by
    filename. The canonical identity is ``(job.date, customer.name)``
    after a relaxed normalization that strips punctuation and
    whitespace so e.g. ``"Dave Atch"`` matches a file whose name is
    just ``"Atch"``.
    """
    index: set[tuple[str, str]] = set()
    for row in list_jobs(jobs_root):
        job = row.get("job") or {}
        customer = row.get("customer") or {}
        date_iso = str(job.get("date") or "")[:10]
        if not date_iso:
            continue
        name_norm = _name_key(customer.get("name"))
        if name_norm:
            index.add((date_iso, name_norm))
            # Also index the last-name-only variant so the dashboard's
            # historic slug convention still matches even when the
            # xlsx supplies a full name.
            last = name_norm.rsplit(" ", 1)[-1]
            if last and last != name_norm:
                index.add((date_iso, last))
    return index


def _name_key(value: object) -> str:
    """Lowercased, punctuation-stripped, single-space-collapsed form of
    a customer name. Used purely as a dedupe key."""
    text = _norm_text(value).lower()
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_phone(value: object) -> str:
    """Render a phone number cell consistently.

    Excel often stores raw 10-digit phones as floats; convert to
    ``(NNN) NNN-NNNN``. Strings that already carry separators are kept
    as-is. Free-text cells like ``"N/A"`` or ``"Facebook"`` (a real
    quirk in this sheet) are dropped — they would mislead the
    dashboard.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        digits = re.sub(r"\D", "", f"{value:.0f}" if isinstance(value, float) else str(value))
    else:
        digits = re.sub(r"\D", "", str(value))
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits[0] == "1":
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    text = _norm_text(value)
    # Drop obvious non-phone free-text. Anything else (e.g. an already-
    # formatted "(440) 985-8801") passes through verbatim.
    if text.lower() in {"n/a", "na", "none", ""} or "@" in text:
        return ""
    if text.isalpha():
        return ""
    return text


def _norm_lead_source(value: object) -> str:
    key = _norm_text(value).lower()
    if not key:
        return ""
    return _LEAD_SOURCE_MAP.get(key, "other")


def _norm_status(value: object) -> str:
    key = _norm_text(value).lower()
    if not key:
        return "booked"
    return _STATUS_MAP.get(key, "booked")


def _norm_payment_method(value: object) -> str | None:
    key = _norm_text(value).lower()
    if not key:
        return None
    return _PAYMENT_MAP.get(key, "other")


def _norm_service_tags(value: object) -> list[str]:
    """Split a free-text service-type cell into canonical tags.

    Combined services like ``"Driveway+ Walkways"`` or ``"Driveway +
    Patio"`` emit two tags. Unknown free-text falls back to ``other``.
    """
    raw = _norm_text(value).lower()
    if not raw:
        return []
    # Common multi-service phrasings — split on + and /.
    parts = [p.strip() for p in re.split(r"[+/]", raw) if p.strip()]
    if not parts:
        return [_SERVICE_TAG_MAP.get(raw, ["other"])[0]]
    tags: list[str] = []
    for part in parts:
        if part in _SERVICE_TAG_MAP:
            tags.extend(_SERVICE_TAG_MAP[part])
            continue
        # "walkways" / "walkway" → front_walk
        if "walk" in part:
            tags.append("front_walk")
            continue
        # Heuristic by substring.
        matched = False
        for needle, mapped in _SERVICE_TAG_MAP.items():
            if needle in part:
                tags.extend(mapped)
                matched = True
                break
        if not matched:
            tags.append("other")
    # Dedupe while preserving order.
    seen: set[str] = set()
    out: list[str] = []
    for t in tags:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def _norm_date(value: object) -> _date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    return None


def _norm_money(value: object) -> float | int | None:
    if value is None or value == "":
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    return int(f) if f.is_integer() else f


def _row_to_payload(row: dict[str, Any]) -> dict[str, Any] | None:
    """Convert one xlsx row dict → bpw_jobs save_job payload, or None
    if the row is unusable (no name AND no date).
    """
    name = _norm_text(row["Name"])
    job_date = _norm_date(row["Job Date"]) or _norm_date(row["Date Contacted"])
    if not name and not job_date:
        return None

    status = _norm_status(row["Status"])
    lead = _norm_lead_source(row["Lead Source"])
    tags_kinds = _norm_service_tags(row["Service Type"])
    payment_method = _norm_payment_method(row["Payment Method"])
    paid = _norm_text(row["Paid?"]).lower() == "yes"
    total = _norm_money(row["Quote $"])
    referred_by = _norm_text(row["Referred By"])
    notes = _norm_text(row["Notes"])
    phone = _norm_phone(row["Phone"])
    address = _norm_text(row["Address"])

    customer: dict[str, Any] = {
        "name": name,
        "phone": phone,
        "address": address,
    }
    if lead:
        customer["lead_source"] = lead
        if lead == "repeat_customer":
            customer["is_repeat"] = True
    if referred_by and referred_by.lower() not in {"n/a", "na", "new lead", "flier"}:
        customer["referred_by"] = referred_by

    payload: dict[str, Any] = {
        "job": {
            "date": job_date.isoformat() if job_date else _date.today().isoformat(),
            "status": status,
        },
        "customer": customer,
        "home": {},
        "tags": [{"type": k, "price": 0} for k in tags_kinds],
        "pricing": {
            "total": total,
            "paid": paid,
            "method": payment_method,
            "discount": "",
        },
        "notes": notes,
    }
    return payload


def _iter_rows(xlsx_path: Path):
    """Yield (row_index, dict_keyed_by_header)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required: pip install openpyxl"
        ) from exc
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Leads & Jobs"]
    header: list[str] | None = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header is None:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        yield i, dict(zip(header, row, strict=False))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument(
        "--jobs-root", type=Path, default=Path(_DEFAULT_BPW_JOBS_ROOT)
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Show which rows would be added without writing.",
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Print every skipped row, not just summary counts.",
    )
    args = parser.parse_args(argv)

    if not args.xlsx.exists():
        print(f"ERROR: xlsx not found at {args.xlsx}", file=sys.stderr)
        return 2
    if not args.jobs_root.exists():
        print(
            f"ERROR: jobs root not found at {args.jobs_root}", file=sys.stderr,
        )
        return 2

    existing = _existing_index(args.jobs_root)

    n_total = 0
    n_skipped_empty = 0
    n_skipped_existing = 0
    n_skipped_invalid = 0
    n_added = 0
    added_files: list[str] = []

    for row_idx, row in _iter_rows(args.xlsx):
        n_total += 1
        try:
            payload = _row_to_payload(row)
        except KeyError as exc:
            print(f"row {row_idx}: missing column {exc}", file=sys.stderr)
            n_skipped_invalid += 1
            continue
        if payload is None:
            n_skipped_empty += 1
            if args.verbose:
                print(f"row {row_idx}: skipped (no name and no date)")
            continue
        # Dedup by (date, name) read from existing YAML content — the
        # on-disk filename convention varies, so a filename check alone
        # would create duplicates.
        date_iso = payload["job"]["date"]
        name_norm = _name_key(payload["customer"].get("name"))
        dedup_keys = {(date_iso, name_norm)}
        if " " in name_norm:
            dedup_keys.add((date_iso, name_norm.rsplit(" ", 1)[-1]))
        if dedup_keys & existing:
            n_skipped_existing += 1
            if args.verbose:
                print(f"row {row_idx}: already on disk (matched by date+name)")
            continue
        target_name = _filename_for(payload)
        target_path = args.jobs_root / target_name
        if target_path.exists():
            # Different person same day same slug — unlikely but guard.
            n_skipped_existing += 1
            if args.verbose:
                print(f"row {row_idx}: filename collision → {target_name}")
            continue
        if args.dry_run:
            print(f"row {row_idx}: WOULD add → {target_name} "
                  f"(name={payload['customer'].get('name')!r}, "
                  f"date={payload['job']['date']}, "
                  f"status={payload['job']['status']}, "
                  f"total={payload['pricing']['total']})")
            n_added += 1
            added_files.append(target_name)
            # Keep the in-memory index honest so two xlsx rows pointing
            # at the same (date, name) don't both pass dedupe.
            existing.update(dedup_keys)
            continue
        try:
            saved = save_job(payload, jobs_root=args.jobs_root)
        except Exception as exc:
            print(f"row {row_idx}: SAVE FAILED → {exc}", file=sys.stderr)
            n_skipped_invalid += 1
            continue
        n_added += 1
        added_files.append(saved["_source_file"])
        existing.update(dedup_keys)
        print(f"row {row_idx}: added → {saved['_source_file']} "
              f"(id={saved['job']['id']})")

    print()
    print(f"xlsx rows scanned       : {n_total}")
    print(f"  empty (no name+date)  : {n_skipped_empty}")
    print(f"  already on disk       : {n_skipped_existing}")
    print(f"  failed to import      : {n_skipped_invalid}")
    print(f"  added{' (dry-run)' if args.dry_run else '              '}: {n_added}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
